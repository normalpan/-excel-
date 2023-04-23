#打开文件夹，指定图片大小（按像素) ,指定存放路径，指定生成的文件名
#-*- coding:utf-8 -*-
import easygui
import time
import os
import shutil
import tkinter as tk #要使用，先导入
from tkinter import *
from tkinter import messagebox
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from PIL import Image as Image2
from PIL import JpegImagePlugin
JpegImagePlugin._getmp = lambda  x:None

gl_del_pat=''

def print_hint(context): #打印格式
    text1.insert(INSERT,'************************************************************\n')            
    text1.insert(INSERT,context)    
    text1.see("end")  
    text1.update()     

def mycopyfile(srcfile,dstpath):     #复制文件
    #print(srcfile,dstpath)                  # 复制函数
    if  os.path.isfile(srcfile):
        fpath,fname=os.path.split(srcfile)             # 分离文件名和路径
        shutil.copy(srcfile, dstpath + '\\'+fname)          # 复制文件
        #print ("mycopyfile copy %s -> %s"%(srcfile, dstpath + '\\'+fname))

def get_outfile(infile, outfile): #重命名压缩文件的名字
    if outfile:
         return outfile
    dir, suffix = os.path.splitext(infile)
    outfile = '{}{}'.format(dir, suffix)
    return outfile

def compress_image(infile, mb,outfile='', step=10, quality=90):#压缩图片大小
    """不改变图片尺寸压缩到指定大小
    :param infile: 压缩源文件
    :param outfile: 压缩文件保存地址
    :param mb: 压缩目标，KB
    :param step: 每次调整的压缩比率
    :param quality: 初始压缩比率
    :return: 压缩文件地址，压缩文件大小
    """
    o_size = os.path.getsize(infile)/1024
    if o_size <= mb:
        return infile
    outfile = get_outfile(infile, outfile)
    while o_size > mb:
        #print('o_size=',o_size,'    mb=',mb)
        im = Image2.open(infile)
        im.save(outfile, quality=quality)
        if quality - step < 0:
            break
        quality -= step
        o_size = os.path.getsize(outfile)/1024
    return outfile, os.path.getsize(outfile)

selected_dict={'file_path':""} #用来存储函数中需返回的值 选定的文件夹
result_dict = {'file_path':""} #用来存储函数中需返回的值 选定的目标文件夹

def convert_path(path: str) -> str: #标准化路径
    return path.replace(r'\/'.replace(os.sep, ''), os.sep)
    
def selectfilepath():#选择文件夹
     sPath = easygui.diropenbox()    
     selected_dict['file_path'] = sPath
     label_2 = tk.Label(frame1, font=('微软雅黑',8, 'bold',),fg='#FF7F50',bg='#F5FFFA',justify='left',text="源文件选定的文件夹:"+selected_dict['file_path'] )
     label_2.place(relx=0, rely=0.6)
     
def resultfilepath():#选定目标文件夹
     sPath = easygui.diropenbox()    
     result_dict['file_path'] = sPath  
     label_2 = tk.Label(frame3, font=('微软雅黑',8, 'bold'),fg='#FF7F50',bg='#F0F8FF',justify='left',text="目标文件名"+result_dict['file_path'])
     label_2.place(relx=0, rely=0.6)
     
def quitapp():#退出程序
     sys.exit()

def startrun():
    #获取源文件夹名称  目标文件夹
    print_hint('开始执行文件!\n')
    src_path=selected_dict['file_path']
    dest_path=result_dict['file_path']
    pic_w=int(entry1.get())
    pic_h=int(entry2.get())
    pic_size=int(entry3.get())
    dest_filename=dest_path+'\图片导入'+str(time.time())+'.xlsx'
    #新建一个目录，复制所有需要处理的图片到这个目录，压缩新建文件夹的所有文件，压缩到300以下。最终根据新建的文件夹处理插入文件
    back_path=src_path+'\\'+str(time.time())
    global gl_del_path
    gl_del_path=back_path
    os.makedirs(back_path)
      
    print_hint('开始复制图片!\n')    
    file_name_list = os.listdir(src_path)
    for srcfile in file_name_list:   
        mycopyfile(src_path+'\\'+srcfile, back_path)       #复制所有照片到待处理文件夹
        print_hint('复制图片：'+srcfile+' 成功!\n')    
    print_hint('图片复制结束!\n')    
    
    file_name_list = os.listdir(back_path)
    
    print_hint('开始压缩图片!\n')             
    for srcfile in file_name_list:      
        compress_image(back_path+'\\'+srcfile,pic_size) #压缩待处理文件夹的所有图片
        print_hint('压缩图片：'+srcfile+' 成功!\n')      
        
    print_hint('图片压缩完毕!\n')    
              
    wb = Workbook()
    ws = wb.active
    #初始化表头
    ws.cell(row=1, column=1).value ='sn'
    ws.cell(row=1, column=2).value ='文件名'
    ws.cell(row=1, column=3).value ='图片'
           
    print_hint('开始插入图片!\n')    
    
    file_name_list = os.listdir(back_path) #处理压缩好的文件
    #计数
    count=0
    for srcfile in file_name_list:     
        count=count+1
        
    src_path=back_path
    sn=2         

    for srcfile in file_name_list:     
        #插入三列 第一列sn 第二列 文件名 低三列图片
        try:     
            image_file=convert_path(src_path+'\\'+srcfile)
            img = Image(image_file)
            
            image_column='C'  
            img.width, img.height = ((pic_w), (pic_h))
            ws.column_dimensions[image_column].width = (pic_w)*0.15
            ws.row_dimensions[sn].height = (pic_h)*0.8
            ws.cell(row=sn, column=1).value =sn-1
            ws.cell(row=sn, column=2).value =srcfile
            ws.add_image(img, anchor=image_column + str(sn)) 
            print_hint('进度：'+str(sn-1)+' / '+str(count)+' 。 插入图片：'+srcfile+' 成功!\n')      
            sn=sn+1
        except:
            print_hint('进度：'+str(sn-1)+' / '+str(count)+'    '+image_file+'不是图片文件！\n')    
            sn=sn+1
            continue
                   
    print_hint('图片插入结束!\n')     
    print_hint('开始保存文件!\n')    
    b21.config(state='norma')
    wb.save(dest_filename)
    wb.close()
    print_hint('文件保存结束!\n文件存放在'+dest_filename)    
    messagebox.showinfo('运行成功','妹子你真好看！')
    return     
    
def del_comp_file():#删除文件夹内的文件 和文件夹 
    #print('待删除目录：'+gl_del_path)
    print_hint('开始清理过程文件:'+gl_del_path+'\n')
    shutil.rmtree(gl_del_path, ignore_errors=True)
    print_hint('清理完毕！\n')
    b21.config(state='disabled')
    
#######################################################
#####################    以下为主程序    #####################
#######################################################

if __name__ == '__main__':     
    win = tk.Tk()#创建一个窗口，因为后面还要用到所以用window这个变量来赋值，可以自行更改
    sw = win.winfo_screenwidth()     #得到屏幕宽度
    sh = win.winfo_screenheight()    #得到屏幕高度
    ww = 500
    wh = 500
    #窗口宽高为500
    x = (sw-ww) / 2
    y = (sh-wh) / 2
    win.geometry("%dx%d+%d+%d" %(ww,wh,x,y))
    win.title("妹子批量插图专用！")
    win.resizable(False, False) #True 窗口可拉升
    # 定义第一个容器 选择文件夹，使用 labelanchor ='w' 来设置标题的方位
    frame1 = tk.LabelFrame(win, text="第一步", fg='#8A2BE2',font=('微软雅黑',9, 'bold'),labelanchor="n",bg='#F5FFFA')
    frame1.place(relx=0, rely=0, relwidth=1, relheight=0.2)
    # 使用 place 控制 LabelFrame 的位置
    label_1 = tk.Label(frame1, font=('微软雅黑',10, 'bold '),bg='#F5FFFA',justify='left',text="选择需要插入图片的文件夹")
    label_1.place(relx=0.1, rely=0.15)
    b = tk.Button(frame1, text="照片存放文件夹",font=('微软雅黑',10, 'bold'),fg='#FFFFFF', bg='#20B2AA', justify='center',command=selectfilepath)
    b.place(relx=0.5, rely=0.15)
    
    # 定义第二个容器 定义插入图片的大小，使用 labelanchor ='w' 来设置标题的方位
    frame2 = tk.LabelFrame(win, text="第二步", fg='#8A2BE2',font=('微软雅黑',9, 'bold'),labelanchor="n",bg='#F0FFFF')
    # 使用 place 控制 LabelFrame 的位置
    frame2.place(relx=0, rely=0.2, relwidth=1, relheight=0.2)
    
    label_1 = tk.Label(frame2, font=('微软雅黑',10, 'bold '),bg='#F0FFFF',justify='left',text="输入插入图片的宽度(像素)")
    label_1.place(relx=0.1, rely=0)
    
    label_2 = tk.Label(frame2, font=('微软雅黑',10, 'bold '),bg='#F0FFFF',justify='left',text="输入插入图片的高度(像素)")
    label_2.place(relx=0.1, rely=0.3)
    
    label_3 = tk.Label(frame2, font=('微软雅黑',10, 'bold '),bg='#F0FFFF',justify='left',text="输入插入图片限定大小(KB)")
    label_3.place(relx=0.1, rely=0.6)
    
    entry1 = tk.Entry(frame2,font=('微软雅黑',9, 'bold'),fg='#20B2AA', bg='#FFFFFF')
    entry1.place(relx=0.5, rely=0)
    entry1.delete(0, "end")
    entry1.insert(0,'150')
    
    entry2 = tk.Entry(frame2,font=('微软雅黑',9, 'bold'),fg='#20B2AA', bg='#FFFFFF')
    entry2.place(relx=0.5, rely=0.3)
    entry2.delete(0, "end")
    entry2.insert(0,'150')
    
    entry3 = tk.Entry(frame2,font=('微软雅黑',9, 'bold'),fg='#20B2AA', bg='#FFFFFF')
    entry3.place(relx=0.5, rely=0.6)
    entry3.delete(0, "end")
    entry3.insert(0,'1500')
    
    # 定义第三个容器，选择存放的路径 使用 labelanchor ='w' 来设置标题的方位
    frame3 = tk.LabelFrame(win, text="第三步", fg='#8A2BE2',font=('微软雅黑',9, 'bold'),labelanchor="n",bg='#F0F8FF')
    # 使用 place 控制 LabelFrame 的位置
    frame3.place(relx=0, rely=0.4, relwidth=1, relheight=0.2)
    
    label_1 = tk.Label(frame3, font=('微软雅黑',10, 'bold '),bg='#F0F8FF',justify='left',text="目标文件夹")
    label_1.place(relx=0.1, rely=0.2)
    b = tk.Button(frame3, text="EXCEL文件存放文件夹", font=('微软雅黑',10, 'bold'),fg='#FFFFFF', bg='#20B2AA',justify='center',command=resultfilepath)
    b.place(relx=0.5, rely=0.15)
    
    # 定义第四个容器，输出结果提示 使用 labelanchor ='w' 来设置标题的方位
     
    frame4 = tk.LabelFrame(win, text="输出结果", fg='#8A2BE2', font=('微软雅黑',9, 'bold'),labelanchor="n",bg='#F8F8FF')
    # 使用 place 控制 LabelFrame 的位置
    frame4.place(relx=0, rely=0.6, relwidth=1, relheight=0.3)
    
    #label_1 = tk.Label(frame4, font=('雅黑',10, 'bold '),bg='#AACDAA',justify='left',text="输出结果")
    #label_1.place(relx=0, rely=0.2)
    
    text1 = tk.Text(frame4, width=55, height=9, undo=True, autoseparators=True,font=('微软雅黑',8, 'normal'),fg='#20B2AA', bg='#FFFFFF')
    text1.place(relx=0, rely=0,relheight=1,relwidth=1)
    
    scroll = tk.Scrollbar(text1,orient=VERTICAL)
    # 放到窗口的右侧, 填充Y竖直方向
    scroll.pack(side=tk.RIGHT,fill=tk.Y)
    # 两个控件关联
    scroll.config(command=text1.yview,bg='#FFFFF0')
    text1.config(yscrollcommand=scroll.set)
    
    # 定义第5个容器，命令区域 使用 labelanchor ='w' 来设置标题的方位
    frame5 = tk.LabelFrame(win, text="命令区",fg='#8A2BE2',font=('微软雅黑',9, 'bold'), labelanchor="n",bg='#FFFFF0')
    # 使用 place 控制 LabelFrame 的位置
    frame5.place(relx=0, rely=0.9, relwidth=1, relheight=0.1)
    #开始插入
    b2 = tk.Button(frame5, text="开    始",font=('微软雅黑',10, 'bold'),fg='#FFFFFF', bg='#20B2AA',justify='center',command=startrun)
    b2.place(relx=0.075, rely=0, relheight=0.9,relwidth=0.25)
    
    b21= tk.Button(frame5, text="删除临时文件",font=('微软雅黑',8, 'bold'),fg='#FF4500', bg='#FFD700',justify='center',command=del_comp_file)
    b21.place(relx=0.375, rely=0,relheight=0.9,relwidth=0.25)
    b21.config(state='disabled')
    #结束运行
    b3= tk.Button(frame5, text="退    出",font=('微软雅黑',10, 'bold'),fg='#FFFFFF', bg='#B22222',justify='center',command=quitapp)
    b3.place(relx=0.675, rely=0,relheight=0.9,relwidth=0.25)
    
    #程序说明
    print_hint("""说明：\n本程序可以实现选定目录，将目录中的图片插入一个新建的excel中。
同时可以设定插入图片的外观大小和存储大小（压缩图片暂只对jpg文件有效)，
所有的压缩文件在选定的目录中单独有个文件夹,不要压缩文件的可以删除临时文件\n""")
    print_hint('')             
    win.mainloop()

